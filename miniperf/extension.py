import logging
import os
import time
import traceback
import shutil
import sys
import threading
import datetime
import time
import glob
import wcwidth
import subprocess
from openpyxl import load_workbook

ROOT_DIR = os.path.abspath(os.path.dirname(__file__))
python_path = os.path.join(os.path.split(sys.executable)[0], "python.exe")

g_webview = None


class Header(object):
    def __init__(self, _comment, _filed, _type, _subtype):
        print("__HEADER__", _comment, _filed, _type, _subtype)
        _map = {"int": "0", "string": ""}
        self.Comment = _comment.replace("\n", '')
        self.Filed = _filed
        self.FiledType = _type
        self.Default = _map.get(self.FiledType, "")
        self.SubType = _subtype


def registered_webview(webview):
    global g_webview
    g_webview = webview


def mkdir_if_not_exists(path, *paths):
    dirname = os.path.join(path, *paths)
    if not os.path.exists(dirname):
        os.mkdir(dirname)
    return dirname


def external_data_dir():
    return mkdir_if_not_exists(ROOT_DIR, ".data")


def post_message_to_js(msg, type):
    global g_webview
    if g_webview:
        g_webview.send_message_to_js(
            {"type": "on_message", "data": {"type": type, "msg": msg}})


def select_app(window):
    files = window.show_file_dialog_for_file("选择要监控的应用", "应用程序(*.exe)")
    if files and len(files) > 0:
        try:
            file = files[0]
            logging.info("import file %s", file)
            return {"ok": True, "msg": file}
        except Exception as err:
            print('Exception', err)
            traceback.print_exc()
            return {"ok": False, "msg": "无法选择应用"}
    return {"ok": False, "msg": "未选择文件"}

def get_svn_status(xlsxpath):
    ret = {}
    output = subprocess.check_output([get_tpsvn(),'status',xlsxpath]).decode().replace('\r', '').replace(' ', '')
    lines = output.split('\n')
    for line in lines:
        status = line[:1]
        name = line[1:]
        # ret.append({'status': line[:1], 'name': line[1:]})
        ret[name] = status
    return ret

def rpc_get_file_list(xlsxpath):
    file_list = []
    try:
        files_status = get_svn_status(xlsxpath)
        print(files_status)

        files = glob.glob(xlsxpath + '/**/*.xls*', recursive=True)
        for f in files:
            if not os.path.basename(f).startswith('~'):
                status = files_status.get(f, 'o')
                file_list.append({'name': os.path.basename(f), 'path': f, 'status': status})

        
        # list = output.replace(' ', '')
        # list = list[2:]
        # list = list.split(r'\n')        
        # for f in list:
        #     name = f[1:-2].split(r'\\')[-1]
        #     if 'xls' in name and not name.startswith('~'):
        #         file_list.append({'name': name, 'path': f[1:-2].replace('\\\\','\\'), 'status':f[0]})
    except Exception as e:
        print(e)
        files = glob.glob(xlsxpath + '/**/*.xls*', recursive=True)
        for f in files:
            if not os.path.basename(f).startswith('~'):
                file_list.append({'name': os.path.basename(f), 'path': f, 'status':'o'})
    # print(xlsxpath)
    # files = glob.glob(xlsxpath + '/**/*.xls*', recursive=True)
    # for f in files:
    #     if not os.path.basename(f).startswith('~'):
    #             file_list.append({'name': os.path.basename(f), 'path': f})
    print(file_list)
    return file_list


def get_cstemplate_file():
    return os.path.join(ROOT_DIR, "asset", "template", "CSTemplate.cs")

def get_tpsvn():
    return os.path.join(ROOT_DIR, "asset", "tpsvn", "svn.exe")

def rpc_gen_cscode(xlsxpath, codepath):
    if not os.path.exists(xlsxpath):
        return {"ok": False, "msg": f"未找到对应文件/{xlsxpath}"}

    try:
        wb = load_workbook(filename=xlsxpath)
        ws = wb.active

        row_count = ws.max_row
        if row_count < 4:
            return {"ok": False, "msg": "配置表有异常,Row < 4"}

        tab_name = os.path.basename(xlsxpath).split(".")[0]
        code_full_path = f"{codepath}/Tab{tab_name}.cs"

        comments = ws[1]

        fileds = []
        template_filed = '\tpublic __FILED__ { get; private set; } // __COMMENT__\n'
        for i in range(0, len(comments)):     
            header = Header(_comment=ws[3][i].value, _subtype=ws[2][i].value, _filed=ws[4][i].value, _type=ws[1][i].value, )       
            line = template_filed.replace(
                "__FILED__", f"{header.FiledType} {header.Filed}")
            line = line.replace("__COMMENT__", f"{header.Comment}")
            fileds.append(line)

        header_list = []
        fileds_parser = []
        template_filed = '\t\t__FILED__ = Get___TYPE__(cellStrs[__INDEX__], "");\n'
        for i in range(0, len(comments)):
            header = Header(_comment=ws[3][i].value, _subtype=ws[2][i].value, _filed=ws[4][i].value, _type=ws[1][i].value, )       
            header_list.append(header)
            if not header.Filed:
                return {"ok": False, "msg": f"[2]{i} is None"}

            line = template_filed.replace("__FILED__", header.Filed)
            line = line.replace("__TYPE__", header.FiledType)
            line = line.replace("__INDEX__", f"{i}")
            fileds_parser.append(line)

        mark_pkey = []
        for e in header_list:
            if e.SubType and 'PKEY' in e.SubType:
                k = e.SubType.split('|')[0]
                t = e.FiledType
                n = e.Filed
                mark_pkey.append((k, t, n))

        with open(get_cstemplate_file(), 'r', encoding="utf-8") as f:
            dat = f.read()
            dat = dat.replace("<TABNAME>", tab_name)
            dat = dat.replace("<TABPATH>", f'"{tab_name}.csv"')
            dat = dat.replace("//MARK_FILEDS", ''.join(fileds))
            dat = dat.replace("//MARK_PARSER", ''.join(fileds_parser))
            for (k, t, n) in mark_pkey:
                dat = dat.replace(f"//MARK_T_{k}", t)
                dat = dat.replace(f"//MARK_N_{k}", n)

            if len(mark_pkey) == 1:
                dat = dat.replace(f"//1st_dec ", '')
                dat = dat.replace(f"//1st_add ", '')
            elif len(mark_pkey) == 2:
                dat = dat.replace(f"//2nd_dec ", '')
                dat = dat.replace(f"//2nd_add ", '')

            with open(code_full_path, 'wb') as f2:
                f2.write(dat.encode("utf-8"))
        return {"ok": True, "dat": f"生成成功,{code_full_path}"}
    except Exception as err:
        print('Exception', err)
        traceback.print_exc()
        return {"ok": False, "msg": "应用产生异常,请联系开发人员"}


def rpc_gen_mgrcode(xlsxpath, mgrpath):
    if not os.path.exists(xlsxpath) or not os.path.exists(mgrpath):
        return {"ok": False, "msg": f"未找到对应文件/{xlsxpath} or {mgrpath}"}

    try:
        tab_name = os.path.basename(xlsxpath).split(".")[0]
        tab_class_name = f'Table{tab_name}'
        tab_var_name = f'Table_{tab_name}'
        dat = ''
        with open(mgrpath, 'r', encoding="utf-8") as f:
            dat = f.read()

        init_block = f"//{tab_class_name}-S-Init\n\t\t{tab_var_name} = new {tab_class_name}();\n\t\t{tab_var_name}.Init();\n\t\t//{tab_class_name}-E-Init\n\t\t//MARK_Init"
        ondestory_block = f"//{tab_class_name}-S-OnDestory\n\t\t{tab_var_name} = null;\n\t\t//{tab_class_name}-E-OnDestory\n\t\t//MARK_OnDestory"
        filed_block = f"public {tab_class_name} {tab_var_name};\n\t//Mark_Fileds"
        if f"//{tab_class_name}-S-Init" not in dat:
            dat = dat.replace("//MARK_Init", init_block)
            dat = dat.replace("//MARK_OnDestory", ondestory_block)
            dat = dat.replace("//Mark_Fileds", filed_block)

        with open(mgrpath, 'wb') as f2:
            f2.write(dat.encode("utf-8"))
        return {"ok": True, "dat": f"生成成功,{mgrpath}"}
    except Exception as err:
        print('Exception', err)
        traceback.print_exc()
        return {"ok": False, "msg": "应用产生异常,请联系开发人员"}

def dump(xlsxpath):
    def wstring(s,width):        
        count = wcwidth.wcswidth(s)-len(s)
        width = width - count if width >= count else 0        
        return '{0:{1}{2}{3}}'.format(s,'','^',width)      

    wb = load_workbook(filename=xlsxpath)
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        cells = [wstring(f"{e.value}", 20).replace("\n", "") for e in ws[i]]
        print('|'.join(cells))


def rpc_gen_tabfile(xlsxpath, tabdir):
    try:
        if not os.path.exists(xlsxpath):
            return {"ok": False, "msg": f"未找到对应文件/{xlsxpath}"}
        if not os.path.exists(tabdir):
            return {"ok": False, "msg": f"未找到对应文件/{tabdir}"}
        dump(xlsxpath)
        
        wb = load_workbook(filename=xlsxpath)
        ws = wb.active
        row_count = ws.max_row
        if row_count < 4:
            return {"ok": False, "msg": "配置表有异常,Row < 4"}

        tab_name = os.path.basename(xlsxpath).split(".")[0]
        tab_fullpath = f"{tabdir}/{tab_name}.csv"
        tab_header = None
        for i in range(0, len(ws[1])):
            tab_header = Header(_comment=ws[3][i].value, _subtype=ws[2][i].value, _filed=ws[4][i].value, _type=ws[1][i].value, )

        with open(tab_fullpath, 'wb') as f:
            for i in range(3, ws.max_row + 1):
                cells = [f'{e.value}' for e in ws[i]]
                for j in range(len(cells)):
                    if cells[j] == 'None':
                        cells[j] = ''  # tab_header[j].Default
                line = ','.join(cells)
                if i != ws.max_row:
                    line += '\n'
                f.write(line.encode("utf-8"))
        return {"ok": True, "dat": f"生成成功,{tab_fullpath}"}

    except Exception as err:
        print('Exception', err)
        traceback.print_exc()
        return {"ok": False, "msg": "应用产生异常,请联系开发人员"}


def rpc_open_folder(dir):
    if '.cs' in dir:
        dir = os.path.dirname(dir)
    open_dir(dir)
    return {"ok": True, "dat": f"{dir}"}


def rpc_get_default_setting(app):
    xlsx_path = app.get_config("Settting/XLSX_PATH", os.path.join(".", "configs.ini"))
    cscode_path = app.get_config("Settting/CSCODE_PATH", os.path.join(".", "configs.ini"))
    csmgr_path = app.get_config("Settting/MGR_PATH", os.path.join(".", "configs.ini"))
    tab_path = app.get_config("Settting/TAB_PATH", os.path.join(".", "configs.ini"))
    dat = {
        "XLSX_PATH": xlsx_path,
        "CSCODE_PATH": cscode_path,
        "TAB_PATH": tab_path,
        "MGR_PATH": csmgr_path
    }
    return {"ok": True, "dat": dat}


def return_log():
    pass


def wait_app_end(process, mtPath, symbol_paths):
    process.wait()
    logging.info("应用监控结束,开始分析数据")
    post_message_to_js("End application", "out")
    analyze_data(mtPath, symbol_paths)


def start(exe_path, symbol_paths=None):
    if not os.path.exists(exe_path) or os.path.isdir(exe_path):
        return {"ok": False, "msg": "选择程序未找到"}
    try:
        data_dir = external_data_dir()
        tmp_dir = mkdir_if_not_exists(os.path.join(data_dir, "tmp"))

        elevatePath = os.path.join(
            ROOT_DIR, ".asset", "memory-trace-toolkit", "Elevate.exe")
        mmtracerPath = os.path.join(
            ROOT_DIR, ".asset", "memory-trace-toolkit", "mmtracer.exe")
        outPath = os.path.join(tmp_dir, "tmp.mt")
        cwd = os.path.split(exe_path)[0]
        process = mkmt(elevatePath, mmtracerPath, outPath, exe_path, cwd)
        post_message_to_js("Start application", "out")
        wait_app_end_thread = threading.Thread(target=wait_app_end, args=(
            process, outPath, symbol_paths), name="wait_app_end_thread")
        wait_app_end_thread.start()
        return {"ok": True, "msg": "启动启动成功"}
    except Exception as err:
        print('Exception', err)
        traceback.print_exc()
        return {"ok": False, "msg": "启动应用失败"}


def analyze_data(mtPath, symbol_paths):
    data_dir = external_data_dir()
    report_dir = mkdir_if_not_exists(data_dir, "report")
    outPath = os.path.join(report_dir, "%s.tpd" % datetime.datetime.fromtimestamp(
        int(time.time())).strftime('%Y_%m_%d_%H_%M_%S'))
    scriptPath = os.path.join(
        ROOT_DIR, ".asset", "memory-trace-toolkit", "scripts", "mktpd.py")
    symbol_path_list = symbol_paths.split(";")
    post_message_to_js("Start analyze", "out")
    process = mktpd(python_path, scriptPath, mtPath,
                    outPath, *symbol_path_list)
    rOut = threading.Thread(target=readOut, args=(process,), name="rOut")
    rErr = threading.Thread(target=readErr, args=(process,), name="rErr")
    rOut.start()
    rErr.start()
    rOut.join()
    rErr.join()
    post_message_to_js("End analyze", "out")
    logging.info("End analyze")
    post_message_to_js("stop", "notice")
    open_dir(report_dir)


def readOut(process):
    logging.info("readOut")
    for readOut_line in process.stdout:
        if len(readOut_line) < 1:
            break
        readOut_line = readOut_line.decode("utf-8")
        logging.info("readOut: %s", readOut_line)
        post_message_to_js(readOut_line, "out")


def readErr(process):
    logging.info("readErr")
    for readErr_line in process.stderr:
        if len(readErr_line) < 1:
            break
        readErr_line = readErr_line.decode("utf-8")
        logging.info("readErr: %s", readErr_line)
        post_message_to_js(readErr_line, "out")


def open_dir(report_dir):
    logging.info("open_dir")
    os.startfile(report_dir)
